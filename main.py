from openpyxl import workbook, load_workbook

wb = load_workbook('employeedata.xlsx')
ws = wb.active
for i in range(2, ws.max_row + 1):
    cell = ws.cell(i, 2)
    if ('helpinghands.cm' in str(cell.value)):
        updated_email = (cell.value).replace('helpinghands.cm', 'handsinhands.org')
        ws.cell(i, 2).value = updated_email
wb.save('newdata.xlsx')
